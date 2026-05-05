"""
多测试者对比分析模块

支持多个测试者的并行对比分析和可视化
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import numpy as np

from .models import AnalysisResult, AnalysisSummary


@dataclass
class TestSubject:
    """测试者信息"""
    name: str
    result: AnalysisResult
    video_path: str | None = None
    analysis_time: str | None = None


@dataclass
class MultiTestComparison:
    """多测试者对比结果"""
    template_name: str
    template_baseline: dict[str, Any]
    subjects: list[TestSubject] = field(default_factory=list)
    comparison_matrix: dict[str, Any] = field(default_factory=dict)
    ranking: list[dict[str, Any]] = field(default_factory=list)


class MultiTestAnalyzer:
    """多测试者分析器"""

    def __init__(self, template_baseline: dict[str, Any]):
        """
        初始化多测试者分析器

        Args:
            template_baseline: 模板基线数据
        """
        self.template_baseline = template_baseline
        self.subjects: list[TestSubject] = []

    def add_subject(
        self,
        name: str,
        result: AnalysisResult,
        video_path: str | None = None,
    ) -> None:
        """
        添加测试者

        Args:
            name: 测试者名称
            result: 分析结果
            video_path: 视频路径
        """
        subject = TestSubject(
            name=name,
            result=result,
            video_path=video_path,
            analysis_time=result.summary.分析时间 if hasattr(result.summary, '分析时间') else None,
        )
        self.subjects.append(subject)

    def compute_comparison_matrix(self) -> dict[str, Any]:
        """
        计算对比矩阵

        Returns:
            对比矩阵数据
        """
        if not self.subjects:
            return {}

        # 提取所有测试者的指标
        metrics = {
            '综合评分': [],
            '手臂动作': [],
            '下盘稳定': [],
            '动作连贯': [],
            '节奏控制': [],
            '时序相似度': [],
        }

        subject_names = []

        for subject in self.subjects:
            subject_names.append(subject.name)
            summary = subject.result.summary

            metrics['综合评分'].append(summary.综合评分 or 0)
            metrics['手臂动作'].append(summary.手臂动作 or 0)
            metrics['下盘稳定'].append(summary.下盘稳定 or 0)
            metrics['动作连贯'].append(summary.动作连贯 or 0)
            metrics['节奏控制'].append(summary.节奏控制 or 0)
            metrics['时序相似度'].append(summary.时序相似度 or 0)

        # 计算统计信息
        comparison_matrix = {
            'subject_names': subject_names,
            'metrics': {},
        }

        for metric_name, values in metrics.items():
            comparison_matrix['metrics'][metric_name] = {
                'values': values,
                'mean': float(np.mean(values)),
                'std': float(np.std(values)),
                'min': float(np.min(values)),
                'max': float(np.max(values)),
                'best_subject': subject_names[np.argmax(values)],
                'worst_subject': subject_names[np.argmin(values)],
            }

        return comparison_matrix

    def compute_ranking(self) -> list[dict[str, Any]]:
        """
        计算排名

        Returns:
            排名列表
        """
        if not self.subjects:
            return []

        # 按综合评分排序
        sorted_subjects = sorted(
            self.subjects,
            key=lambda s: s.result.summary.综合评分 or 0,
            reverse=True
        )

        ranking = []
        for rank, subject in enumerate(sorted_subjects, start=1):
            summary = subject.result.summary
            ranking.append({
                'rank': rank,
                'name': subject.name,
                'score': summary.综合评分 or 0,
                'grade': self._get_grade(summary.综合评分 or 0),
                'issues_count': len(subject.result.issues),
                'critical_issues': sum(1 for issue in subject.result.issues if issue.severity == 'critical'),
            })

        return ranking

    def analyze_common_issues(self) -> dict[str, Any]:
        """
        分析共性问题

        Returns:
            共性问题分析结果
        """
        if not self.subjects:
            return {}

        # 统计所有问题类型
        issue_categories = {}

        for subject in self.subjects:
            for issue in subject.result.issues:
                category = issue.category
                if category not in issue_categories:
                    issue_categories[category] = {
                        'count': 0,
                        'subjects': [],
                        'severity_counts': {'critical': 0, 'major': 0, 'minor': 0},
                    }

                issue_categories[category]['count'] += 1
                if subject.name not in issue_categories[category]['subjects']:
                    issue_categories[category]['subjects'].append(subject.name)

                severity = issue.severity or 'minor'
                issue_categories[category]['severity_counts'][severity] += 1

        # 找出共性问题（超过50%的测试者都有的问题）
        threshold = len(self.subjects) * 0.5
        common_issues = []

        for category, data in issue_categories.items():
            if len(data['subjects']) >= threshold:
                common_issues.append({
                    'category': category,
                    'frequency': len(data['subjects']) / len(self.subjects),
                    'subjects': data['subjects'],
                    'total_count': data['count'],
                    'severity_distribution': data['severity_counts'],
                })

        # 按频率排序
        common_issues.sort(key=lambda x: x['frequency'], reverse=True)

        return {
            'common_issues': common_issues,
            'total_issue_types': len(issue_categories),
            'total_subjects': len(self.subjects),
        }

    def generate_comparison_report(self) -> MultiTestComparison:
        """
        生成对比报告

        Returns:
            多测试者对比结果
        """
        comparison_matrix = self.compute_comparison_matrix()
        ranking = self.compute_ranking()

        return MultiTestComparison(
            template_name=self.template_baseline.get('模板名称', '未命名模板'),
            template_baseline=self.template_baseline,
            subjects=self.subjects,
            comparison_matrix=comparison_matrix,
            ranking=ranking,
        )

    def export_comparison_excel(self, output_path: Path) -> None:
        """
        导出对比Excel报表

        Args:
            output_path: 输出路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()

        # 1. 综合排名表
        ws_ranking = wb.active
        ws_ranking.title = "综合排名"

        headers = ['排名', '姓名', '综合评分', '评价等级', '问题数量', '严重问题']
        ws_ranking.append(headers)

        # 设置表头样式
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_ranking.cell(1, col_idx)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 填充数据
        ranking = self.compute_ranking()
        for rank_data in ranking:
            ws_ranking.append([
                rank_data['rank'],
                rank_data['name'],
                round(rank_data['score'], 1),
                rank_data['grade'],
                rank_data['issues_count'],
                rank_data['critical_issues'],
            ])

        # 2. 详细指标对比表
        ws_metrics = wb.create_sheet("详细指标对比")

        comparison_matrix = self.compute_comparison_matrix()
        subject_names = comparison_matrix.get('subject_names', [])

        # 表头
        headers = ['指标'] + subject_names + ['平均值', '最高分', '最低分']
        ws_metrics.append(headers)

        # 设置表头样式
        for col_idx in range(1, len(headers) + 1):
            cell = ws_metrics.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 填充指标数据
        for metric_name, metric_data in comparison_matrix.get('metrics', {}).items():
            row = [metric_name]
            row.extend([round(v, 1) for v in metric_data['values']])
            row.append(round(metric_data['mean'], 1))
            row.append(round(metric_data['max'], 1))
            row.append(round(metric_data['min'], 1))
            ws_metrics.append(row)

        # 3. 共性问题分析表
        ws_issues = wb.create_sheet("共性问题分析")

        headers = ['问题类别', '出现频率', '涉及人数', '总次数', '严重', '一般', '轻微']
        ws_issues.append(headers)

        # 设置表头样式
        for col_idx in range(1, len(headers) + 1):
            cell = ws_issues.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.font = Font(bold=True, color='000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 填充共性问题数据
        common_issues_data = self.analyze_common_issues()
        for issue in common_issues_data.get('common_issues', []):
            ws_issues.append([
                issue['category'],
                f"{issue['frequency'] * 100:.0f}%",
                len(issue['subjects']),
                issue['total_count'],
                issue['severity_distribution']['critical'],
                issue['severity_distribution']['major'],
                issue['severity_distribution']['minor'],
            ])

        # 调整列宽
        for ws in [ws_ranking, ws_metrics, ws_issues]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # 保存
        wb.save(output_path)

    def _get_grade(self, score: float) -> str:
        """获取评价等级"""
        if score >= 90:
            return "优秀"
        elif score >= 75:
            return "良好"
        elif score >= 60:
            return "及格"
        else:
            return "需改进"


def create_comparison_visualization(
    comparison: MultiTestComparison,
    output_dir: Path,
) -> dict[str, Path]:
    """
    创建对比可视化图表

    Args:
        comparison: 对比结果
        output_dir: 输出目录

    Returns:
        生成的图表文件路径
    """
    try:
        import matplotlib.pyplot as plt
        import matplotlib
        matplotlib.use('Agg')
    except ImportError:
        raise ImportError("需要安装 matplotlib: pip install matplotlib")

    output_dir.mkdir(parents=True, exist_ok=True)
    generated_files = {}

    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False

    # 1. 综合评分柱状图
    fig, ax = plt.subplots(figsize=(12, 6))

    names = [s.name for s in comparison.subjects]
    scores = [s.result.summary.综合评分 or 0 for s in comparison.subjects]

    colors = ['#4472C4' if s >= 90 else '#70AD47' if s >= 75 else '#FFC000' if s >= 60 else '#C00000' for s in scores]

    bars = ax.bar(names, scores, color=colors, alpha=0.8, edgecolor='black', linewidth=1.5)

    # 在柱子上显示数值
    for bar, score in zip(bars, scores):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width() / 2., height,
                f'{score:.1f}',
                ha='center', va='bottom', fontsize=11, fontweight='bold')

    ax.set_xlabel('测试者', fontsize=12, fontweight='bold')
    ax.set_ylabel('综合评分', fontsize=12, fontweight='bold')
    ax.set_title('多测试者综合评分对比', fontsize=14, fontweight='bold', pad=20)
    ax.set_ylim(0, 105)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    score_chart_path = output_dir / '01_综合评分对比.png'
    plt.savefig(score_chart_path, dpi=150, bbox_inches='tight')
    plt.close()
    generated_files['score_chart'] = score_chart_path

    # 2. 雷达图（多维度对比）
    if len(comparison.subjects) <= 5:  # 雷达图最多显示5个测试者
        fig, ax = plt.subplots(figsize=(10, 10), subplot_kw=dict(projection='polar'))

        categories = ['手臂动作', '下盘稳定', '动作连贯', '节奏控制', '时序相似度']
        num_vars = len(categories)

        angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
        angles += angles[:1]

        ax.set_theta_offset(np.pi / 2)
        ax.set_theta_direction(-1)
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(categories, fontsize=11)
        ax.set_ylim(0, 100)

        colors_radar = ['#4472C4', '#70AD47', '#FFC000', '#C00000', '#7030A0']

        for idx, subject in enumerate(comparison.subjects):
            summary = subject.result.summary
            values = [
                summary.手臂动作 or 0,
                summary.下盘稳定 or 0,
                summary.动作连贯 or 0,
                summary.节奏控制 or 0,
                summary.时序相似度 or 0,
            ]
            values += values[:1]

            ax.plot(angles, values, 'o-', linewidth=2, label=subject.name,
                    color=colors_radar[idx % len(colors_radar)])
            ax.fill(angles, values, alpha=0.15, color=colors_radar[idx % len(colors_radar)])

        ax.set_title('多维度能力雷达图', fontsize=14, fontweight='bold', pad=20)
        ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=10)
        ax.grid(True)

        plt.tight_layout()

        radar_chart_path = output_dir / '02_多维度雷达图.png'
        plt.savefig(radar_chart_path, dpi=150, bbox_inches='tight')
        plt.close()
        generated_files['radar_chart'] = radar_chart_path

    return generated_files
