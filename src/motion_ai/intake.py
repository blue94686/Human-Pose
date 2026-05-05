from __future__ import annotations

import json
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from openpyxl import load_workbook

from .config import DEFAULT_TEMPLATE_FILE, DEFAULT_WEIGHTS
from .description import DescriptionParser
from .pose import COCO_KEYPOINTS, PoseEstimator, PoseFrame


IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".bmp", ".webp"}
TEXT_SUFFIXES = {".txt", ".md", ".json"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}

_BODY_PART_LABELS = {
    "arms": "上肢",
    "torso": "躯干",
    "legs": "下肢",
    "balance": "重心",
    "tempo": "节奏",
}

_TEMPO_LABELS = {
    "slow": "节奏缓慢",
    "medium": "节奏平稳",
    "fast": "节奏较快",
}


@dataclass
class IntakeSuggestion:
    source_type: str
    source_path: str = ""
    title: str = ""
    extracted_text: str = ""
    suggested_description: str = ""
    template_name: str | None = None
    keywords: list[str] = field(default_factory=list)
    focus_body_parts: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    details: dict[str, str] = field(default_factory=dict)

    def to_report_text(self) -> str:
        lines: list[str] = []
        if self.title:
            lines.append(f"标题：{self.title}")
        lines.append(f"来源：{self.source_type}")
        if self.source_path:
            lines.append(f"路径：{self.source_path}")
        if self.template_name:
            lines.append(f"匹配模板：{self.template_name}")
        if self.keywords:
            lines.append(f"关键词：{'、'.join(self.keywords)}")
        if self.focus_body_parts:
            labels = [_BODY_PART_LABELS.get(part, part) for part in self.focus_body_parts]
            lines.append(f"关注部位：{'、'.join(labels)}")
        if self.suggested_description:
            lines.append("")
            lines.append("建议录入描述：")
            lines.append(self.suggested_description)
        if self.details:
            lines.append("")
            lines.append("提取信息：")
            for key, value in self.details.items():
                lines.append(f"- {key}：{value}")
        if self.notes:
            lines.append("")
            lines.append("说明：")
            for note in self.notes:
                lines.append(f"- {note}")
        if self.extracted_text:
            lines.append("")
            lines.append("原始/摘要内容：")
            lines.append(self.extracted_text)
        return "\n".join(lines).strip()


def analyze_text_source(path_text: str | Path, template_file: str | Path | None = None) -> IntakeSuggestion:
    path = Path(path_text).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{path}")

    parser = DescriptionParser(Path(template_file) if template_file else DEFAULT_TEMPLATE_FILE)
    suffix = path.suffix.lower()
    raw_text = path.read_text(encoding="utf-8")
    condensed = _condense_text(raw_text)
    notes = [
        f"文件类型：{suffix or '无扩展名'}",
        f"字符数：{len(raw_text)}",
    ]

    title = path.name
    extracted_text = condensed
    suggested_description = condensed

    if suffix == ".json":
        extracted_text, suggested_description, extra_notes = _summarize_json_file(path, raw_text, parser)
        notes.extend(extra_notes)
    elif not suggested_description:
        suggested_description = _suggest_from_filename(path.stem, parser)

    if not suggested_description:
        suggested_description = "健身气功动作录入，建议补充动作名称、节奏、躯干姿态和重点关注部位。"

    rules = parser.parse(suggested_description)
    if not rules.template_name:
        filename_template = _match_template_from_name(path.stem, parser)
        if filename_template:
            rules = parser.parse(_compose_description_from_template(filename_template, parser))

    return IntakeSuggestion(
        source_type="文本/模板文件",
        source_path=str(path),
        title=title,
        extracted_text=extracted_text,
        suggested_description=suggested_description,
        template_name=rules.template_name,
        keywords=rules.keywords,
        focus_body_parts=rules.focus_body_parts,
        notes=notes,
        details={
            "建议节奏": _TEMPO_LABELS.get(rules.expected_tempo, rules.expected_tempo),
            "对称要求": "需要" if rules.requires_symmetry else "一般",
            "躯干要求": "保持直立" if rules.requires_upright_torso else "按动作自定",
            "定势要求": "包含停顿/保持" if rules.requires_hold else "连续动作优先",
        },
    )


def analyze_text_content(text: str, template_file: str | Path | None = None, title: str = "当前描述文本") -> IntakeSuggestion:
    parser = DescriptionParser(Path(template_file) if template_file else DEFAULT_TEMPLATE_FILE)
    content = (text or "").strip()
    if not content:
        raise ValueError("当前没有可解析的文本内容。")
    condensed = _condense_text(content)
    rules = parser.parse(condensed)
    return IntakeSuggestion(
        source_type="描述文本",
        title=title,
        extracted_text=condensed,
        suggested_description=condensed,
        template_name=rules.template_name,
        keywords=rules.keywords,
        focus_body_parts=rules.focus_body_parts,
        notes=[f"字符数：{len(content)}"],
        details={
            "建议节奏": _TEMPO_LABELS.get(rules.expected_tempo, rules.expected_tempo),
            "对称要求": "需要" if rules.requires_symmetry else "一般",
            "躯干要求": "保持直立" if rules.requires_upright_torso else "按动作自定",
            "定势要求": "包含停顿/保持" if rules.requires_hold else "连续动作优先",
        },
    )


def analyze_image_source(
    path_text: str | Path,
    weights_path: str | Path | None = None,
    template_file: str | Path | None = None,
) -> tuple[IntakeSuggestion, np.ndarray]:
    path = Path(path_text).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"图片不存在：{path}")

    frame = cv2.imread(str(path))
    if frame is None:
        raise RuntimeError(f"无法读取图片：{path}")

    parser = DescriptionParser(Path(template_file) if template_file else DEFAULT_TEMPLATE_FILE)
    notes = [
        "静态图片适合录入姿态信息，不适合直接判断完整节奏和连续性。",
        f"图像尺寸：{frame.shape[1]} x {frame.shape[0]}",
    ]
    details: dict[str, str] = {}

    template_hint = _match_template_from_name(path.stem, parser)
    pose_estimator = PoseEstimator(Path(weights_path) if weights_path else DEFAULT_WEIGHTS)
    pose = pose_estimator.estimate(frame) if pose_estimator.available else None

    description = ""
    if pose is not None:
        description, details, pose_notes = _describe_pose(pose)
        notes.extend(pose_notes)
    else:
        notes.append(pose_estimator.status.reason)
        notes.append("已完成图片读取和预览，可在放入姿态权重后生成更细的骨架录入建议。")
        description = "健身气功动作图片录入，建议补充动作名称、左右协调情况和躯干要求。"

    if template_hint:
        notes.insert(0, f"文件名命中模板提示：{template_hint}")
        description = _merge_template_hint(description, template_hint)

    rules = parser.parse(description)
    if template_hint and not rules.template_name:
        rules = parser.parse(_compose_description_from_template(template_hint, parser))

    return (
        IntakeSuggestion(
            source_type="图片/抓拍",
            source_path=str(path),
            title=path.name,
            extracted_text=f"已读取图片：{path.name}\n尺寸：{frame.shape[1]} x {frame.shape[0]}",
            suggested_description=description,
            template_name=rules.template_name or template_hint,
            keywords=rules.keywords,
            focus_body_parts=rules.focus_body_parts,
            notes=notes,
            details=details,
        ),
        frame,
    )


def analyze_spreadsheet_source(path_text: str | Path, template_file: str | Path | None = None) -> IntakeSuggestion:
    path = Path(path_text).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"表格不存在：{path}")

    parser = DescriptionParser(Path(template_file) if template_file else DEFAULT_TEMPLATE_FILE)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        preview_lines: list[str] = []
        notes: list[str] = [f"表格工作表数：{len(workbook.sheetnames)}"]
        collected_text: list[str] = []

        for sheet_name in workbook.sheetnames[:4]:
            sheet = workbook[sheet_name]
            row_count = 0
            preview_lines.append(f"[工作表] {sheet_name}")
            for row in sheet.iter_rows(min_row=1, max_row=12, values_only=True):
                cells = [str(cell).strip() for cell in row if cell not in (None, "")]
                if not cells:
                    continue
                row_count += 1
                line = " | ".join(cells[:6])
                preview_lines.append(f"- {line}")
                collected_text.append(" ".join(cells))
                if row_count >= 5:
                    break
            if row_count == 0:
                preview_lines.append("- 空表或前 12 行无有效内容")
    finally:
        workbook.close()

    raw_text = "\n".join(collected_text)
    condensed = _condense_text("\n".join(preview_lines), max_lines=18, max_chars=900)
    suggested_description = _condense_text(raw_text, max_lines=10, max_chars=300)
    if not suggested_description:
        suggested_description = _suggest_from_filename(path.stem, parser)
    if not suggested_description:
        suggested_description = "动作表格已导入，建议补充动作名称、阶段节奏、重点部位和判定标准。"

    rules = parser.parse(suggested_description)
    if not rules.template_name:
        filename_template = _match_template_from_name(path.stem, parser)
        if filename_template:
            rules = parser.parse(_compose_description_from_template(filename_template, parser))

    notes.append("适合导入动作步骤、指标阈值、评分字段等结构化模板资料。")
    return IntakeSuggestion(
        source_type="Excel 表格",
        source_path=str(path),
        title=path.name,
        extracted_text=condensed,
        suggested_description=suggested_description,
        template_name=rules.template_name,
        keywords=rules.keywords,
        focus_body_parts=rules.focus_body_parts,
        notes=notes,
        details={
            "建议节奏": _TEMPO_LABELS.get(rules.expected_tempo, rules.expected_tempo),
            "对称要求": "需要" if rules.requires_symmetry else "一般",
            "躯干要求": "保持直立" if rules.requires_upright_torso else "按动作自定",
            "定势要求": "包含停顿/保持" if rules.requires_hold else "连续动作优先",
        },
    )


def _summarize_json_file(
    path: Path,
    raw_text: str,
    parser: DescriptionParser,
) -> tuple[str, str, list[str]]:
    notes: list[str] = []
    try:
        payload = json.loads(raw_text)
    except json.JSONDecodeError:
        notes.append("JSON 解析失败，已按普通文本读取。")
        condensed = _condense_text(raw_text)
        return condensed, condensed, notes

    if isinstance(payload, dict) and all(isinstance(value, dict) for value in payload.values()):
        names = list(payload.keys())
        preview_lines = [f"模板数量：{len(names)}", f"模板名称：{'、'.join(names[:6])}"]
        example_name = names[0] if names else ""
        suggested = _compose_description_from_template(example_name, parser) if example_name else ""
        notes.append("检测到模板库 JSON，已转成可直接录入的动作描述建议。")
        return "\n".join(preview_lines), suggested, notes

    condensed = _condense_text(json.dumps(payload, ensure_ascii=False, indent=2))
    notes.append("检测到普通 JSON 数据，已提取摘要内容。")
    return condensed, condensed, notes


def _describe_pose(pose: PoseFrame) -> tuple[str, dict[str, str], list[str]]:
    notes: list[str] = ["已检测到单人骨架，可用于录入姿态信息。"]
    details: dict[str, str] = {}

    left_shoulder = _point(pose, "left_shoulder")
    right_shoulder = _point(pose, "right_shoulder")
    left_wrist = _point(pose, "left_wrist")
    right_wrist = _point(pose, "right_wrist")
    left_hip = _point(pose, "left_hip")
    right_hip = _point(pose, "right_hip")
    left_elbow = _point(pose, "left_elbow")
    right_elbow = _point(pose, "right_elbow")
    left_knee = _point(pose, "left_knee")
    right_knee = _point(pose, "right_knee")
    left_ankle = _point(pose, "left_ankle")
    right_ankle = _point(pose, "right_ankle")
    nose = _point(pose, "nose")

    shoulder_width = max(_distance(left_shoulder, right_shoulder), 1.0)
    shoulder_center = (left_shoulder + right_shoulder) / 2.0
    hip_center = (left_hip + right_hip) / 2.0
    ankle_distance = _distance(left_ankle, right_ankle)

    left_raise = left_shoulder[1] - left_wrist[1]
    right_raise = right_shoulder[1] - right_wrist[1]
    arm_state = "双臂自然下垂"
    if left_raise > shoulder_width * 0.15 and right_raise > shoulder_width * 0.15:
        arm_state = "双臂上举"
    elif left_raise > shoulder_width * 0.15 or right_raise > shoulder_width * 0.15:
        arm_state = "单侧手臂抬起"
    elif left_wrist[1] < left_elbow[1] or right_wrist[1] < right_elbow[1]:
        arm_state = "手腕略高于肘部，处于抬臂准备或过渡姿态"

    wrist_gap = abs(left_wrist[1] - right_wrist[1])
    symmetric = wrist_gap <= shoulder_width * 0.18
    shoulder_gap = abs(left_shoulder[1] - right_shoulder[1])
    hip_gap = abs(left_hip[1] - right_hip[1])
    torso_tilt_deg = _torso_tilt_deg(shoulder_center, hip_center)

    stance_ratio = ankle_distance / shoulder_width
    stance_text = "站姿较稳"
    if stance_ratio >= 1.35:
        stance_text = "步幅较开"
    elif stance_ratio <= 0.8:
        stance_text = "步幅较窄"

    head_relation = "头部位于双肩之间"
    if abs(nose[0] - shoulder_center[0]) > shoulder_width * 0.3:
        head_relation = "头部相对肩部有偏移"

    torso_text = "躯干基本直立" if abs(torso_tilt_deg) <= 12 else "躯干存在侧倾"
    knee_text = "双膝基本伸直"
    if left_knee[1] < left_hip[1] + shoulder_width * 0.9 or right_knee[1] < right_hip[1] + shoulder_width * 0.9:
        knee_text = "膝部有轻微屈曲"

    correction_parts: list[str] = []
    if not symmetric:
        correction_parts.append("调整两侧手腕高度，让左右手同时到位")
    if abs(torso_tilt_deg) > 12:
        correction_parts.append("收紧核心，使肩髋连线回到垂直中线")
    if stance_ratio <= 0.8:
        correction_parts.append("适当打开步幅，增加支撑稳定性")
    elif stance_ratio >= 1.6:
        correction_parts.append("略收窄步幅，避免重心过度分散")
    if not correction_parts:
        correction_parts.append("保持当前中正站姿，录入视频时补充起势、过渡、定势和收势过程")

    fragments = [
        "图片中为单人站姿",
        arm_state,
        "左右较为对称" if symmetric else "左右存在高低差",
        torso_text,
        stance_text,
        knee_text,
    ]
    description = (
        "，".join(fragments)
        + f"。建议作为录入参考时记录：手臂状态为{arm_state}，躯干倾斜约{torso_tilt_deg:.1f}°，"
        + f"步幅约为肩宽的{stance_ratio:.2f}倍。若测试者动作不一致，优先修改："
        + "；".join(correction_parts)
        + "。"
    )

    details["手臂状态"] = arm_state
    details["对称情况"] = "较对称" if symmetric else "存在偏差"
    details["腕部高度差"] = f"{wrist_gap / shoulder_width:.2f} 个肩宽"
    details["肩部水平差"] = f"{shoulder_gap / shoulder_width:.2f} 个肩宽"
    details["髋部水平差"] = f"{hip_gap / shoulder_width:.2f} 个肩宽"
    details["躯干倾斜"] = f"{torso_tilt_deg:.1f}°"
    details["步幅比例"] = f"{stance_ratio:.2f}"
    details["下肢状态"] = knee_text
    details["头肩关系"] = head_relation
    details["动作修改依据"] = "；".join(correction_parts)

    if left_wrist[1] < nose[1] and right_wrist[1] < nose[1]:
        notes.append("双腕位置高于头部，可能是上举/托举类动作。")
    if not symmetric:
        notes.append("左右腕部高度差较明显，建议结合视频进一步确认动作对称性。")
    if abs(torso_tilt_deg) > 12:
        notes.append("躯干侧倾偏大，建议录入时明确躯干中正要求。")

    return description, details, notes


def _compose_description_from_template(template_name: str, parser: DescriptionParser) -> str:
    template = parser.templates.get(template_name, {})
    full_description = str(template.get("description") or "").strip()
    phases = template.get("phases") or []
    common_errors = template.get("common_errors") or []
    corrections = template.get("corrections") or []
    if full_description:
        lines = [f"{template_name}动作模板：{full_description}"]
        if phases:
            lines.append("动作阶段：" + "；".join(str(item) for item in phases[:8]))
        if common_errors:
            lines.append("常见错误：" + "；".join(str(item) for item in common_errors[:6]))
        if corrections:
            lines.append("纠正建议：" + "；".join(str(item) for item in corrections[:6]))
        return "\n".join(lines)

    tempo = _TEMPO_LABELS.get(str(template.get("tempo", "medium")), "节奏平稳")
    fragments = [f"{template_name}动作", tempo]
    if template.get("requires_symmetry"):
        fragments.append("双侧动作保持对称")
    if template.get("requires_upright_torso"):
        fragments.append("躯干保持直立")
    if template.get("requires_hold"):
        fragments.append("关键姿态注意停顿保持")
    focus_parts = [_BODY_PART_LABELS.get(part, part) for part in template.get("focus_body_parts", [])]
    if focus_parts:
        fragments.append(f"重点关注{'、'.join(focus_parts)}")
    return "，".join(fragments) + "。"


def _merge_template_hint(description: str, template_name: str) -> str:
    if template_name in description:
        return description
    return f"{template_name}动作录入参考，{description}"


def _match_template_from_name(text: str, parser: DescriptionParser) -> str | None:
    for name, template in parser.templates.items():
        if name in text:
            return name
        if any(keyword in text for keyword in template.get("keywords", [])):
            return name
    return None


def _suggest_from_filename(text: str, parser: DescriptionParser) -> str:
    template_name = _match_template_from_name(text, parser)
    if template_name:
        return _compose_description_from_template(template_name, parser)
    return ""


def _condense_text(text: str, max_lines: int = 8, max_chars: int = 420) -> str:
    lines = [line.strip() for line in text.replace("\r", "").split("\n") if line.strip()]
    if not lines:
        return ""
    condensed = "\n".join(lines[:max_lines])
    if len(condensed) > max_chars:
        condensed = condensed[: max_chars - 1].rstrip() + "…"
    return condensed


def _point(pose: PoseFrame, name: str) -> np.ndarray:
    index = COCO_KEYPOINTS.index(name)
    point = pose.keypoints[index]
    return np.array([float(point[0]), float(point[1])], dtype=np.float32)


def _distance(a: np.ndarray, b: np.ndarray) -> float:
    return float(np.linalg.norm(a - b))


def _torso_tilt_deg(shoulder_center: np.ndarray, hip_center: np.ndarray) -> float:
    vector = shoulder_center - hip_center
    if np.linalg.norm(vector) == 0:
        return 0.0
    return math.degrees(math.atan2(float(vector[0]), max(abs(float(vector[1])), 1e-6)))
