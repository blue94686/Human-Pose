from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

from .models import AnalysisResult, FrameMetrics, PoseFrame
from .pose import COCO_KEYPOINTS


_TRACKING_MODE_LABELS = {
    "detector": "模型识别",
    "detector+flow": "模型识别+时序跟踪",
    "flow_fallback": "光流兜底",
    "flow_only": "纯光流跟踪",
    "hold_last": "沿用上一帧",
    "mediapipe": "MediaPipe 实时识别",
}

_KEYPOINT_LABELS = {
    "nose": "鼻尖",
    "left_eye": "左眼",
    "right_eye": "右眼",
    "left_ear": "左耳",
    "right_ear": "右耳",
    "left_shoulder": "左肩",
    "right_shoulder": "右肩",
    "left_elbow": "左肘",
    "right_elbow": "右肘",
    "left_wrist": "左腕",
    "right_wrist": "右腕",
    "left_hip": "左髋",
    "right_hip": "右髋",
    "left_knee": "左膝",
    "right_knee": "右膝",
    "left_ankle": "左踝",
    "right_ankle": "右踝",
}

_STATUS_LABELS = {
    "excellent": "极佳",
    "good": "良好",
    "fair": "一般",
    "warning": "注意",
    "poor": "偏差",
    "missing": "缺失",
    "unknown": "未知",
}


def _sanitize_export_text(value):
    if isinstance(value, dict):
        return {str(k): _sanitize_export_text(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_sanitize_export_text(item) for item in value]
    if isinstance(value, tuple):
        return [_sanitize_export_text(item) for item in value]
    if not isinstance(value, str):
        return value
    text = value
    replacements = {
        "鈿狅笍": "",
        "鈿": "",
        "鉁": "",
        "鈥": "-",
        "**": "",
        "\uFFFD": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return " ".join(text.split())


def export_result_json(result: AnalysisResult, output_path: Path, *, include_frame_rows: bool = True) -> None:
    payload = _sanitize_export_text(build_chinese_result_payload(result, include_frame_rows=include_frame_rows))
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export_comparison_json(result: AnalysisResult, output_path: Path) -> None:
    payload = _sanitize_export_text(build_chinese_comparison_payload(result))
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export_frame_metrics_csv(frame_metrics: list[FrameMetrics], output_path: Path) -> None:
    fieldnames = [
        "frame",
        "timestamp_sec",
        "keypoint",
        "x",
        "y",
        "confidence",
        "motion_mean",
        "left_right_balance",
        "torso_tilt_deg",
        "arm_symmetry_error",
        "joint_jitter",
        "skeleton_stability",
        "template_distance",
    ]
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for item in frame_metrics:
            positions = item.keypoint_positions or {}
            for keypoint_name in COCO_KEYPOINTS:
                coords = positions.get(keypoint_name)
                if coords is None or len(coords) < 2:
                    continue
                writer.writerow(
                    {
                        "frame": item.frame_index,
                        "timestamp_sec": round(float(item.timestamp_sec), 3),
                        "keypoint": keypoint_name,
                        "x": round(float(coords[0]), 3),
                        "y": round(float(coords[1]), 3),
                        "confidence": _estimate_export_confidence(item, keypoint_name),
                        "motion_mean": round(float(item.motion_mean), 4),
                        "left_right_balance": round(float(item.left_right_balance), 4),
                        "torso_tilt_deg": _round_value(item.torso_tilt_deg, 2),
                        "arm_symmetry_error": _round_value(item.arm_symmetry_error, 2),
                        "joint_jitter": _round_value(item.joint_jitter, 4),
                        "skeleton_stability": _round_value(item.skeleton_stability, 4),
                        "template_distance": _round_value(item.template_distance, 6),
                    }
                )


def export_pose_keypoints_csv(
    pose_frames: list[PoseFrame],
    frame_metrics: list[FrameMetrics],
    output_path: Path,
    *,
    is_smoothed: bool,
) -> None:
    fieldnames = [
        "frame_index",
        "timestamp",
        "track_id",
        "keypoint_name",
        "x",
        "y",
        "confidence",
        "is_valid",
        "is_smoothed",
    ]
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for index, metrics in enumerate(frame_metrics):
            pose = pose_frames[index] if index < len(pose_frames) else None
            confidences = pose.keypoint_confidences if pose else []
            keypoints = pose.keypoints if pose else []
            track_id = pose.track_id if pose else None
            for kp_index, keypoint_name in enumerate(COCO_KEYPOINTS):
                x = ""
                y = ""
                confidence = ""
                is_valid = 0
                if pose is not None and kp_index < len(keypoints):
                    point = keypoints[kp_index]
                    point_is_finite = _is_finite_point(point)
                    if point_is_finite:
                        x = round(float(point[0]), 4)
                        y = round(float(point[1]), 4)
                    if kp_index < len(confidences):
                        confidence = round(float(confidences[kp_index]), 4)
                        is_valid = 1 if point_is_finite and float(confidences[kp_index]) >= 0.5 else 0
                    else:
                        is_valid = 1 if point_is_finite else 0
                writer.writerow(
                    {
                        "frame_index": metrics.frame_index,
                        "timestamp": round(float(metrics.timestamp_sec), 4),
                        "track_id": track_id if track_id is not None else "",
                        "keypoint_name": keypoint_name,
                        "x": x,
                        "y": y,
                        "confidence": confidence,
                        "is_valid": is_valid,
                        "is_smoothed": 1 if is_smoothed else 0,
                    }
                )


def export_summary_metrics_json(result: AnalysisResult, output_path: Path) -> None:
    inhibition = result.inhibition_metrics
    payload = {
        "source": result.source,
        "analysis_mode": result.analysis_mode,
        "fps": result.fps,
        "frame_count": result.frame_count,
        "processed_frames": result.processed_frames,
        "used_pose_estimator": result.used_pose_estimator,
        "research_metrics": {
            "Stability_CV": getattr(inhibition, "body_stability_cv", 0.0) if inhibition else 0.0,
            "Transition_Acc_Peak": getattr(inhibition, "transition_acc_peak", 0.0) if inhibition else 0.0,
            "Transition_Jerk_Mean": getattr(inhibition, "transition_jerk_mean", 0.0) if inhibition else 0.0,
            "Control_Ratio": getattr(inhibition, "body_control_ratio", 0.0) if inhibition else 0.0,
        },
        "summary": build_chinese_result_payload(result, include_frame_rows=False),
    }
    output_path.write_text(json.dumps(_sanitize_export_text(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def export_inhibition_metrics_xlsx(result: AnalysisResult, output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "科研指标"
    headers = ["指标类别", "指标名称", "数值", "单位", "说明"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    inhibition = result.inhibition_metrics
    if inhibition is None:
        sheet.append(["状态", "未生成", "", "", "当前分析未产出可用姿态序列。"])
    else:
        rows = [
            ["基础信息", "总帧数", inhibition.frame_count, "帧", "参与统计的视频帧数。"],
            ["基础信息", "有效帧数", inhibition.valid_frame_count, "帧", "关键点满足阈值要求的有效帧数。"],
            ["肢体晃动度", "总体 RMS", inhibition.limb_sway_degree, "px/frame", "四肢末端帧间位移均方根，越小越稳定。"],
            ["躯干稳定度", "中心点时序位移", inhibition.torso_stability_degree, "px", "肩髋中心点位移方差的综合表征，越小越稳。"],
            ["规范偏差", "模板对齐平均距离", inhibition.posture_deviation, "归一化距离", "经过对齐后的平均欧氏距离，越小越贴近标准。"],
            ["无效小动作频次", "计数", inhibition.invalid_motion_count, "次", "定势或低运动阶段内速度超阈值的微动次数。"],
            ["科研指标", "身体稳定性参数（CV）", getattr(inhibition, "body_stability_cv", 0.0), "CV", "颈-髋连线动态偏移角度的变异系数，越小越稳定。"],
            ["科研指标", "动作流畅性参数（过渡期加速度峰值）", getattr(inhibition, "transition_acc_peak", 0.0), "px/frame²", "过渡阶段髋部中心点加速度峰值，越小越流畅。"],
            ["科研指标", "动作流畅性参数（过渡期急动度均值）", getattr(inhibition, "transition_jerk_mean", 0.0), "px/frame³", "过渡阶段髋部中心点急动度均值，越小越平顺。"],
            ["科研指标", "身体调控能力参数", getattr(inhibition, "body_control_ratio", 0.0), "ratio", "过渡阶段位移幅度与定势阶段位移幅度之比，可直接用于后测导出。"],
        ]
        for row in rows:
            sheet.append(row)

        if inhibition.limb_sway_detail:
            sheet.append([])
            sheet.append(["肢体晃动明细", "关键点", "标准差", "RMS", "说明"])
            for name, detail in inhibition.limb_sway_detail.items():
                sheet.append(
                    [
                        "肢体晃动明细",
                        _KEYPOINT_LABELS.get(name, name),
                        detail.get("std", 0.0),
                        detail.get("rms", 0.0),
                        "腕、踝等末端关键点的帧间波动情况。",
                    ]
                )

        if inhibition.joint_angle_stability:
            sheet.append([])
            sheet.append(["角度稳定性", "关节", "方差", "变化幅度", "说明"])
            for name, variance in inhibition.joint_angle_stability.items():
                sheet.append(
                    [
                        "角度稳定性",
                        _KEYPOINT_LABELS.get(name, name),
                        variance,
                        inhibition.joint_angle_variation.get(name, 0.0),
                        "角速度方差越小，关节控制越平稳。",
                    ]
                )

    for column in range(1, 6):
        sheet.column_dimensions[chr(64 + column)].width = [18, 28, 18, 18, 48][column - 1]
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    workbook.save(output_path)


def export_analysis_report_markdown(result: AnalysisResult, output_path: Path) -> None:
    lines = [
        "# 动作分析报告",
        "",
        f"- 分析模式：{result.analysis_mode}",
        f"- 输入源：`{result.source}`",
        f"- 总帧数：{result.frame_count}",
        f"- 实际处理帧数：{result.processed_frames}",
        f"- 姿态后端：{'YOLOv8-Pose' if result.used_pose_estimator else '光流兜底'}",
        "",
        "## 综合评分",
        "",
        f"- 总分：{result.summary.total_score:.1f}",
        f"- 姿态标准度：{result.summary.posture_score:.1f}",
        f"- 动作连续性：{result.summary.continuity_score:.1f}",
        f"- 动作稳定性：{result.summary.stability_score:.1f}",
        f"- 节奏控制：{result.summary.rhythm_score:.1f}",
        f"- 动作完整性：{result.summary.completeness_score:.1f}",
    ]
    if result.inhibition_metrics is not None:
        inhibition = result.inhibition_metrics
        lines.extend(
            [
                "",
                "## 科研指标",
                "",
                f"- 身体稳定性参数（CV）：{getattr(inhibition, 'body_stability_cv', 0.0):.6f}",
                f"- 动作流畅性参数（加速度峰值）：{getattr(inhibition, 'transition_acc_peak', 0.0):.6f}",
                f"- 动作流畅性参数（急动度均值）：{getattr(inhibition, 'transition_jerk_mean', 0.0):.6f}",
                f"- 身体调控能力参数：{getattr(inhibition, 'body_control_ratio', 0.0):.6f}",
            ]
        )
    lines.extend(["", "## 问题与建议", ""])
    if result.issues:
        for issue in result.issues:
            lines.append(f"- {issue.title}：{issue.detail} 建议：{issue.suggestion}")
    else:
        lines.append("- 当前未检测到明显问题。")
    if result.summary.advice:
        lines.extend(["", "## 综合建议", ""])
        for advice in result.summary.advice:
            lines.append(f"- {advice}")
    output_path.write_text(_sanitize_export_text("\n".join(lines)), encoding="utf-8")


def _is_finite_point(point: list[float] | tuple[float, ...]) -> bool:
    if len(point) < 2:
        return False
    try:
        return point[0] == point[0] and point[1] == point[1]
    except Exception:
        return False


def export_summary_text(text: str, output_path: Path) -> None:
    output_path.write_text(_sanitize_export_text(text), encoding="utf-8")


def export_prepost_summary_xlsx(rows: list[dict], output_path: Path) -> None:
    if pd is None or not rows:
        return
    normalized_rows = [_sanitize_export_text(row) for row in rows]
    raw_df = pd.DataFrame(normalized_rows)
    improvement_rows: list[dict] = []
    if not raw_df.empty and "Subject_ID" in raw_df.columns and "Test_Type" in raw_df.columns:
        for subject_id, group in raw_df.groupby("Subject_ID", dropna=False):
            pre_row = group[group["Test_Type"] == "前测"]
            post_row = group[group["Test_Type"] == "后测"]
            if pre_row.empty or post_row.empty:
                continue
            pre = pre_row.iloc[0]
            post = post_row.iloc[0]
            improvement_rows.append(
                {
                    "Subject_ID": subject_id,
                    "Baseline_Name": post.get("Baseline_Name") or pre.get("Baseline_Name"),
                    "Stability_Improvement": _safe_metric_diff(pre.get("Stability_CV"), post.get("Stability_CV")),
                    "Fluency_Improvement": _safe_metric_diff(pre.get("Transition_Acc_Peak"), post.get("Transition_Acc_Peak")),
                    "Jerk_Improvement": _safe_metric_diff(pre.get("Transition_Jerk_Mean"), post.get("Transition_Jerk_Mean")),
                    "Control_Ratio_Change": _safe_metric_diff(pre.get("Control_Ratio"), post.get("Control_Ratio")),
                }
            )
    improvement_df = pd.DataFrame(improvement_rows)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name="原始指标", index=False)
        if not improvement_df.empty:
            improvement_df.to_excel(writer, sheet_name="改善量汇总", index=False)


def build_prepost_metric_row(
    result: AnalysisResult,
    *,
    subject_id: str,
    test_type: str,
    baseline_name: str = "",
) -> dict:
    inhibition = result.inhibition_metrics
    return {
        "Subject_ID": subject_id,
        "Test_Type": test_type or "未标记",
        "Baseline_Name": baseline_name,
        "Source": result.source,
        "Stability_CV": getattr(inhibition, "body_stability_cv", 0.0) if inhibition else 0.0,
        "Transition_Acc_Peak": getattr(inhibition, "transition_acc_peak", 0.0) if inhibition else 0.0,
        "Transition_Jerk_Mean": getattr(inhibition, "transition_jerk_mean", 0.0) if inhibition else 0.0,
        "Control_Ratio": getattr(inhibition, "body_control_ratio", 0.0) if inhibition else 0.0,
    }


def _safe_metric_diff(before, after) -> float | None:
    try:
        if before is None or after is None:
            return None
        return round(float(before) - float(after), 6)
    except (TypeError, ValueError):
        return None


def export_result_excel(result: AnalysisResult, output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_overview_sheet(workbook, "总览评分", result, include_comparison=result.analysis_mode != "template")
    _write_summary_sheet(workbook, "阶段与问题", result)
    _write_frame_sheet(workbook, "逐帧指标", result.frame_metrics)
    _write_suggestion_sheet(workbook, "建议汇总", result)
    if result.analysis_mode != "template":
        _write_comparison_sheet(workbook, "模板对比", result)
    if result.inhibition_metrics:
        _write_inhibition_sheet(workbook, "科研指标", result)

    workbook.save(output_path)


def export_unified_report(result: AnalysisResult, output_dir: Path) -> dict[str, Path]:
    output_files = {}

    excel_path = output_dir / "分析报告.xlsx"
    export_result_excel(result, excel_path)
    output_files["Excel报告"] = excel_path

    json_path = output_dir / "完整数据.json"
    export_result_json(result, json_path, include_frame_rows=True)
    output_files["JSON数据"] = json_path

    guidance_path = output_dir / "训练指导.txt"
    guidance_content = _build_unified_guidance_text(result)
    guidance_path.write_text(_sanitize_export_text(guidance_content), encoding="utf-8")
    output_files["训练指导"] = guidance_path

    return output_files


def _build_unified_guidance_text(result: AnalysisResult) -> str:
    lines = []
    lines.append("=" * 60)
    lines.append("动作分析训练指导")
    lines.append("=" * 60)
    lines.append("")
    lines.append(f"分析模式: {'模板原始数据生成' if result.analysis_mode == 'template' else '测试动作评估'}")
    lines.append(f"模板名称: {result.rules.template_name or '未匹配模板'}")
    lines.append(f"测试者水平: {result.summary.learner_level}")
    lines.append(f"评价标准: {result.summary.evaluator_level}")
    lines.append("")
    lines.append("-" * 60)
    lines.append("【综合评分】")
    lines.append("-" * 60)
    lines.append(f"总分: {result.summary.total_score:.1f}")
    lines.append(f"姿态标准度: {result.summary.posture_score:.1f}")
    lines.append(f"动作连续性: {result.summary.continuity_score:.1f}")
    lines.append(f"动作稳定性: {result.summary.stability_score:.1f}")
    lines.append(f"节奏控制: {result.summary.rhythm_score:.1f}")
    lines.append(f"动作完整性: {result.summary.completeness_score:.1f}")
    lines.append("")

    if result.summary.phases:
        lines.append("-" * 60)
        lines.append("【动作阶段分析】")
        lines.append("-" * 60)
        for phase in result.summary.phases:
            name = phase.get("name", "未命名阶段")
            start = phase.get("start_sec", 0)
            end = phase.get("end_sec", 0)
            duration = phase.get("duration_sec", 0)
            lines.append(f"{name}: {start:.2f}s - {end:.2f}s (时长 {duration:.2f}s)")
        lines.append("")

    if result.issues:
        lines.append("-" * 60)
        lines.append("【检测到的问题】")
        lines.append("-" * 60)
        for i, issue in enumerate(result.issues, 1):
            lines.append(f"{i}. {issue.title}")
            lines.append(f"   问题说明: {issue.detail}")
            if issue.suggestion:
                lines.append(f"   纠正建议: {issue.suggestion}")
            if issue.time_range:
                lines.append(f"   问题时段: {issue.time_range[0]:.2f}s - {issue.time_range[1]:.2f}s")
            lines.append("")

    if result.summary.advice:
        lines.append("-" * 60)
        lines.append("【改进建议】")
        lines.append("-" * 60)
        for i, advice in enumerate(result.summary.advice, 1):
            lines.append(f"{i}. {advice}")
        lines.append("")

    if result.summary.improvement_focus:
        lines.append("-" * 60)
        lines.append("【优化方向】")
        lines.append("-" * 60)
        for i, focus in enumerate(result.summary.improvement_focus, 1):
            lines.append(f"{i}. {focus}")
        lines.append("")

    if result.comparison:
        lines.append("-" * 60)
        lines.append("【模板对比】")
        lines.append("-" * 60)
        lines.append(f"基线名称: {result.comparison.baseline_name}")
        lines.append(f"时序相似度: {result.comparison.sequence_similarity_score:.1f} ({result.comparison.sequence_similarity_label})")
        lines.append(f"总体判断: {result.comparison.overall_assessment}")
        lines.append("")
        if result.comparison.metrics:
            lines.append("指标对比:")
            for metric in result.comparison.metrics:
                lines.append(
                    f"  - {metric.label}: 标准 {metric.baseline_value}{metric.unit} / 当前 {metric.current_value}{metric.unit} / {_status_label(metric.status)}"
                )
        lines.append("")

    lines.append("=" * 60)
    lines.append("建议按以上问题与建议分阶段训练，再进行完整复测。")
    lines.append("=" * 60)
    return "\n".join(lines)


def build_chinese_result_payload(result: AnalysisResult, *, include_frame_rows: bool = True) -> dict:
    frame_rows = [build_chinese_frame_metric_row(item) for item in result.frame_metrics] if include_frame_rows else []
    output_files = {
        "综合数据": result.artifacts.get("compact_data_json"),
        "基础报告": result.artifacts.get("excel_report"),
        "结果摘要": result.artifacts.get("analysis_summary_txt"),
        "训练指导": result.artifacts.get("action_guidance_txt"),
        "模板基线数据": result.artifacts.get("template_baseline_json"),
    }
    comparison_payload = None
    if result.comparison is not None:
        comparison_payload = {
            "基线名称": result.comparison.baseline_name,
            "基线来源": result.comparison.baseline_source,
            "基线类型": result.comparison.baseline_kind,
            "对齐方式": result.comparison.alignment_mode,
            "时序相似度评分": result.comparison.sequence_similarity_score,
            "时序相似度结论": result.comparison.sequence_similarity_label,
            "总体判断": result.comparison.overall_assessment,
            "当前缺口": result.comparison.missing_items,
            "对比快照": result.comparison.snapshot_pairs,
            "指标对比": [
                {
                    "指标": item.label,
                    "标准值": item.baseline_value,
                    "当前值": item.current_value,
                    "差值": item.delta_value,
                    "单位": item.unit,
                    "状态": _status_label(item.status),
                    "说明": item.note,
                }
                for item in result.comparison.metrics
            ],
            "阶段对比": [
                {
                    "标准阶段": item.baseline_name,
                    "当前阶段": item.current_name,
                    "标准时长秒": item.baseline_duration_sec,
                    "当前时长秒": item.current_duration_sec,
                    "时长差值秒": item.delta_duration_sec,
                    "说明": item.note,
                }
                for item in result.comparison.phases
            ],
        }
    return {
        "数据分层": {
            "基础报告": "Excel 文件，集中展示总览、阶段、问题、逐帧数据和模板对比。",
            "综合数据": "JSON 文件，包含逐帧记录、基础指标、问题建议和模板对比。",
            "结果摘要": "TXT 文件，便于直接查阅与提交。",
            "训练指导": "TXT 文件，单独保存训练提示、修改建议和优化方向。",
            "模板基线数据": "模板模式生成的 JSON，可供后续测试者动作比对复用。",
        },
        "分析模式": result.analysis_mode,
        "输入源": result.source,
        "视频帧率": result.fps,
        "视频总帧数": result.frame_count,
        "实际处理帧数": result.processed_frames,
        "姿态识别已启用": result.used_pose_estimator,
        "模板名称": result.rules.template_name or "未匹配模板",
        "规则信息": {
            "原始描述": result.rules.raw_text,
            "节奏要求": result.rules.expected_tempo,
            "需要对称": result.rules.requires_symmetry,
            "需要躯干中正": result.rules.requires_upright_torso,
            "需要定势停顿": result.rules.requires_hold,
            "关注部位": result.rules.focus_body_parts,
            "关键词": result.rules.keywords,
            "动作类别": result.rules.action_category,
            "阈值": result.rules.thresholds,
        },
        "综合摘要": {
            "总分": result.summary.total_score,
            "评价标准": result.summary.evaluator_level,
            "测试者水平": result.summary.learner_level,
            "姿态标准度": result.summary.posture_score,
            "动作连续性": result.summary.continuity_score,
            "动作稳定性": result.summary.stability_score,
            "节奏控制": result.summary.rhythm_score,
            "动作完整性": result.summary.completeness_score,
            "平均运动强度": result.summary.avg_motion,
            "平均左右平衡差": result.summary.avg_left_right_balance,
            "平均躯干倾斜角度": result.summary.avg_torso_tilt,
            "平均双臂对称误差": result.summary.avg_arm_symmetry_error,
            "平均关键点覆盖率": result.summary.avg_keypoint_coverage,
            "平均姿态质量分": result.summary.avg_pose_quality_score,
            "姿态质量等级": result.summary.pose_quality_level,
        },
        "动作阶段": result.summary.phases,
        "优化方向": result.summary.improvement_focus,
        "问题列表": [
            {
                "问题编码": issue.code,
                "严重程度": issue.severity,
                "错误类型": issue.error_type,
                "问题标题": issue.title,
                "问题说明": issue.detail,
                "纠正建议": issue.suggestion,
                "关键指标": issue.highlight_metric,
                "预期值": issue.expected_value,
                "实际值": issue.actual_value,
                "画面标记": issue.overlay_label,
                "问题时间区间": list(issue.time_range) if issue.time_range else None,
            }
            for issue in result.issues
        ],
        "建议列表": result.summary.advice,
        "详细对比": comparison_payload,
        "输出文件": output_files,
        "科研后测运动学指标": (
            {
                "身体稳定性参数（CV）": getattr(result.inhibition_metrics, "body_stability_cv", 0.0),
                "动作流畅性参数-过渡期加速度峰值": getattr(result.inhibition_metrics, "transition_acc_peak", 0.0),
                "动作流畅性参数-过渡期急动度均值": getattr(result.inhibition_metrics, "transition_jerk_mean", 0.0),
                "身体调控能力参数": getattr(result.inhibition_metrics, "body_control_ratio", 0.0),
                "说明": "单次结果仅导出前测或后测原始值；改善量需基于前测减后测另行汇总。",
            }
            if result.inhibition_metrics is not None
            else None
        ),
        "逐帧数据": frame_rows if include_frame_rows else None,
    }


def build_chinese_comparison_payload(result: AnalysisResult) -> dict:
    comparison = result.comparison
    return {
        "输入源": result.source,
        "模板名称": result.rules.template_name or "未匹配模板",
        "姿态识别已启用": result.used_pose_estimator,
        "详细对比": (
            {
                "基线名称": comparison.baseline_name,
                "基线来源": comparison.baseline_source,
                "基线类型": comparison.baseline_kind,
                "对齐方式": comparison.alignment_mode,
                "时序相似度评分": comparison.sequence_similarity_score,
                "时序相似度结论": comparison.sequence_similarity_label,
                "总体判断": comparison.overall_assessment,
                "当前缺口": comparison.missing_items,
                "对比快照": comparison.snapshot_pairs,
                "指标对比": [
                    {
                        "指标": item.label,
                        "标准值": item.baseline_value,
                        "当前值": item.current_value,
                        "差值": item.delta_value,
                        "单位": item.unit,
                        "状态": _status_label(item.status),
                        "说明": item.note,
                    }
                    for item in comparison.metrics
                ],
                "阶段对比": [
                    {
                        "标准阶段": item.baseline_name,
                        "当前阶段": item.current_name,
                        "标准时长秒": item.baseline_duration_sec,
                        "当前时长秒": item.current_duration_sec,
                        "时长差值秒": item.delta_duration_sec,
                        "说明": item.note,
                    }
                    for item in comparison.phases
                ],
            }
            if comparison is not None
            else None
        ),
    }


def build_chinese_frame_metric_row(item: FrameMetrics) -> dict[str, object]:
    return {
        "帧序号": item.frame_index,
        "时间秒": _round_value(item.timestamp_sec),
        "运动强度均值": _round_value(item.motion_mean),
        "运动波动值": _round_value(item.motion_std),
        "左右平衡差": _round_value(item.left_right_balance),
        "运动方向角度": _round_value(item.direction_deg),
        "活动区域占比": _round_ratio(item.active_region_ratio),
        "活动区域框": item.motion_focus_bbox,
        "躯干倾斜角度": _round_value(item.torso_tilt_deg),
        "肩部平衡差": _round_value(item.shoulder_balance),
        "左臂抬起角度": _round_value(item.arm_raise_left_deg),
        "右臂抬起角度": _round_value(item.arm_raise_right_deg),
        "双臂对称误差": _round_value(item.arm_symmetry_error),
        "左腕速度": _round_value(item.wrist_speed_left),
        "右腕速度": _round_value(item.wrist_speed_right),
        "骨架位移": _round_value(item.pose_motion),
        "骨架抖动": _round_value(item.joint_jitter),
        "骨架稳定度": _round_value(item.skeleton_stability),
        "姿态置信度": _round_value(item.pose_confidence),
        "关键点覆盖率": _round_ratio(item.keypoint_coverage),
        "可见关键点数": item.visible_keypoint_count,
        "姿态质量分": _round_value(item.pose_quality_score),
        "姿态质量等级": item.pose_quality_label,
        "衔接评分": _round_value(item.transition_score),
        "时序状态": item.sequence_state,
        "关键点坐标": json.dumps(_translate_keypoint_positions(item.keypoint_positions), ensure_ascii=False),
        "跟踪方式": _tracking_mode_label(item.tracking_mode),
        "动作阶段": item.phase_label,
        "模板偏差距离": _round_value(item.template_distance, 6),
    }


def _round_value(value, digits: int = 4):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), digits)
    return value


def _round_ratio(value):
    if value is None:
        return None
    return round(float(value) * 100.0, 2)


def _tracking_mode_label(value: str | None) -> str | None:
    if value is None:
        return None
    return _TRACKING_MODE_LABELS.get(value, value)


def _translate_keypoint_positions(positions: dict[str, list[float]]) -> dict[str, list[float]]:
    return {_KEYPOINT_LABELS.get(key, key): [round(float(v), 2) for v in value] for key, value in positions.items()}


def _status_label(value: str | None) -> str | None:
    if value is None:
        return None
    return _STATUS_LABELS.get(value.lower(), value) if isinstance(value, str) else value


def _write_overview_sheet(
    workbook: Workbook,
    title: str,
    result: AnalysisResult,
    *,
    include_comparison: bool,
) -> None:
    sheet = workbook.create_sheet(title)
    rows = [
        ("分析模式", "模板原始数据生成" if result.analysis_mode == "template" else "测试者动作分析"),
        ("输入源", result.source),
        ("模板名称", result.rules.template_name or "未匹配模板"),
        ("评价标准", result.summary.evaluator_level),
        ("测试者水平", result.summary.learner_level),
        ("姿态识别", "YOLOv8-Pose 已启用" if result.used_pose_estimator else "未启用，当前使用光流兜底分析"),
        ("实际处理帧数", result.processed_frames),
        ("视频总帧数", result.frame_count or "未知"),
        ("视频帧率", round(float(result.fps), 2)),
        ("总分", round(float(result.summary.total_score), 1)),
        ("姿态标准度", round(float(result.summary.posture_score), 1)),
        ("动作连续性", round(float(result.summary.continuity_score), 1)),
        ("动作稳定性", round(float(result.summary.stability_score), 1)),
        ("节奏控制", round(float(result.summary.rhythm_score), 1)),
        ("动作完整性", round(float(result.summary.completeness_score), 1)),
        ("平均运动强度", _round_value(result.summary.avg_motion)),
        ("平均左右平衡差", _round_value(result.summary.avg_left_right_balance)),
        ("平均躯干倾斜角度", _round_value(result.summary.avg_torso_tilt)),
        ("平均双臂对称误差", _round_value(result.summary.avg_arm_symmetry_error)),
        ("平均抬臂角度", _round_value(result.summary.avg_arm_raise)),
        ("平均关键点覆盖率", _round_ratio(result.summary.avg_keypoint_coverage)),
        ("平均姿态质量分", _round_value(result.summary.avg_pose_quality_score)),
        ("姿态质量等级", result.summary.pose_quality_level),
    ]
    if include_comparison and result.comparison is not None:
        rows.extend(
            [
                ("基线名称", result.comparison.baseline_name),
                ("基线类型", result.comparison.baseline_kind),
                ("对齐方式", result.comparison.alignment_mode),
                ("时序相似度评分", _round_value(result.comparison.sequence_similarity_score)),
                ("时序相似度结论", result.comparison.sequence_similarity_label),
                ("总体判断", result.comparison.overall_assessment),
            ]
        )
    _write_key_value_sheet(sheet, rows)


def _write_summary_sheet(workbook: Workbook, title: str, result: AnalysisResult) -> None:
    sheet = workbook.create_sheet(title)
    rows = [["类型", "名称", "说明/建议", "时间范围"]]

    for phase in result.summary.phases or []:
        rows.append(
            [
                "动作阶段",
                phase.get("name") or "未命名阶段",
                f"{phase.get('start_sec', 0):.2f}s - {phase.get('end_sec', 0):.2f}s，时长 {phase.get('duration_sec', 0):.2f}s",
                f"{phase.get('start_sec', 0):.2f}s - {phase.get('end_sec', 0):.2f}s",
            ]
        )

    for issue in result.issues:
        time_range = ""
        if issue.time_range:
            time_range = f"{issue.time_range[0]:.2f}s - {issue.time_range[1]:.2f}s"
        rows.append(
            [
                issue.error_type or "动作问题",
                issue.title,
                f"{issue.detail} / 指标：{issue.highlight_metric or '未标注'} / 改进：{issue.suggestion}",
                time_range,
            ]
        )

    for advice in result.summary.advice:
        rows.append(["改进建议", "综合建议", advice, ""])
    for item in result.summary.improvement_focus:
        rows.append(["优化方向", result.summary.learner_level or "未评定", item, ""])

    _write_table_sheet(sheet, rows)


def _write_frame_sheet(workbook: Workbook, title: str, frame_metrics: list[FrameMetrics]) -> None:
    sheet = workbook.create_sheet(title)
    rows = [[
        "帧序号",
        "时间秒",
        "运动强度均值",
        "运动波动值",
        "左右平衡差",
        "活动区域占比",
        "躯干倾斜角度",
        "双臂对称误差",
        "骨架抖动",
        "骨架稳定度",
        "姿态置信度",
        "关键点覆盖率",
        "可见关键点数",
        "姿态质量分",
        "姿态质量等级",
        "衔接评分",
        "时序状态",
        "跟踪方式",
        "动作阶段",
        "模板偏差距离",
    ]]
    for item in frame_metrics:
        rows.append(
            [
                item.frame_index,
                _round_value(item.timestamp_sec),
                _round_value(item.motion_mean),
                _round_value(item.motion_std),
                _round_value(item.left_right_balance),
                _round_ratio(item.active_region_ratio),
                _round_value(item.torso_tilt_deg),
                _round_value(item.arm_symmetry_error),
                _round_value(item.joint_jitter),
                _round_value(item.skeleton_stability),
                _round_value(item.pose_confidence),
                _round_ratio(item.keypoint_coverage),
                item.visible_keypoint_count,
                _round_value(item.pose_quality_score),
                item.pose_quality_label,
                _round_value(item.transition_score),
                item.sequence_state,
                _tracking_mode_label(item.tracking_mode),
                item.phase_label,
                _round_value(item.template_distance, 6),
            ]
        )
    _write_table_sheet(sheet, rows)


def _write_suggestion_sheet(workbook: Workbook, title: str, result: AnalysisResult) -> None:
    sheet = workbook.create_sheet(title)
    rows = [["类型", "标题", "内容"]]
    if result.issues:
        for issue in result.issues:
            rows.append(["问题定位", issue.title, issue.detail])
            rows.append(["修改建议", issue.title, issue.suggestion])
    else:
        rows.append(["修改建议", "整体状态", "当前未发现明显异常，建议继续补充更多模板或测试视频进行细粒度比对。"])
    for advice in result.summary.advice:
        rows.append(["综合建议", "训练建议", advice])
    for item in result.summary.improvement_focus:
        rows.append(["优化方向", "重点关注", item])
    _write_table_sheet(sheet, rows)


def _estimate_export_confidence(item: FrameMetrics, _keypoint_name: str) -> float | None:
    if item.pose_confidence is None:
        return None
    return round(float(item.pose_confidence), 4)


def _write_comparison_sheet(workbook: Workbook, title: str, result: AnalysisResult) -> None:
    sheet = workbook.create_sheet(title)
    comparison = result.comparison
    rows = [["类别", "项目", "标准值", "当前值", "差值", "状态/说明"]]
    if comparison is None:
        rows.append(["模板对比", "暂无", "", "", "", "当前没有可导出的详细对比结果"])
        _write_table_sheet(sheet, rows)
        return

    for metric in comparison.metrics:
        rows.append(
            [
                "指标对比",
                metric.label,
                _format_excel_metric_value(metric.baseline_value, metric.unit),
                _format_excel_metric_value(metric.current_value, metric.unit),
                _format_excel_metric_value(metric.delta_value, metric.unit, signed=True),
                _status_label(metric.status or "未知"),
            ]
        )
    for phase in comparison.phases:
        rows.append(
            [
                "阶段对比",
                phase.baseline_name or "未定义",
                _format_excel_seconds(phase.baseline_duration_sec),
                _format_excel_seconds(phase.current_duration_sec),
                _format_excel_seconds(phase.delta_duration_sec, signed=True),
                phase.note,
            ]
        )
    for item in comparison.missing_items:
        rows.append(["当前缺口", "缺口说明", "", "", "", item])
    for snapshot in comparison.snapshot_pairs:
        rows.append(
            [
                "对比快照",
                snapshot.get("phase_name") or "未命名阶段",
                _format_excel_seconds(snapshot.get("baseline_time_sec")),
                _format_excel_seconds(snapshot.get("current_time_sec")),
                "",
                (
                    f"当前运动强度 {snapshot.get('current_motion')} / "
                    f"左右平衡差 {snapshot.get('current_balance')} / "
                    f"姿态质量 {snapshot.get('current_pose_quality')}"
                ),
            ]
        )
    rows.append(["总体判断", "结论", "", "", "", comparison.overall_assessment])
    _write_table_sheet(sheet, rows)


def _write_inhibition_sheet(workbook: Workbook, title: str, result: AnalysisResult) -> None:
    sheet = workbook.create_sheet(title)
    inhibition = result.inhibition_metrics
    if not inhibition:
        sheet.cell(row=1, column=1, value="未计算身体动作抑制指标")
        return

    rows = [
        ["指标类别", "指标名称", "数值", "单位", "说明"],
        ["基础信息", "总帧数", inhibition.frame_count, "帧", ""],
        ["基础信息", "有效帧数", inhibition.valid_frame_count, "帧", ""],
        ["基础信息", "置信度阈值", inhibition.confidence_threshold, "", ""],
        ["时序波动", "躯干中心漂移", f"{inhibition.torso_center_drift:.2f}", "像素", "数值越小，抑制能力越强。"],
    ]

    if inhibition.keypoint_std:
        avg_std = sum(inhibition.keypoint_std.values()) / len(inhibition.keypoint_std)
        rows.append(["时序波动", "关键点平均标准差", f"{avg_std:.2f}", "像素", "各关键点坐标标准差的平均值。"])
    if inhibition.keypoint_fluctuation:
        avg_fluct = sum(inhibition.keypoint_fluctuation.values()) / len(inhibition.keypoint_fluctuation)
        rows.append(["时序波动", "关键点平均波动幅度", f"{avg_fluct:.2f}", "像素", "连续帧间位移的平均值。"])

    for limb_name, dispersion in inhibition.limb_dispersion.items():
        limb_label = {"left_arm": "左臂", "right_arm": "右臂", "left_leg": "左腿", "right_leg": "右腿"}.get(limb_name, limb_name)
        rows.append(["四肢离散度", f"{limb_label}位移离散度", f"{dispersion:.2f}", "像素", "位移标准差。"])

    if inhibition.template_euclidean_distance > 0:
        rows.append(["姿态偏差", "模板欧氏距离", f"{inhibition.template_euclidean_distance:.2f}", "像素", "越小越接近标准动作。"])

    rows.extend(
        [
            ["姿态偏差", "动作规范偏差", f"{inhibition.posture_deviation:.2f}", "像素", "与标准示范的总体偏差。"],
            ["抑制指标", "肢体晃动度", f"{inhibition.limb_sway_degree:.2f}", "像素", "四肢关键点帧间波动值。"],
            ["抑制指标", "躯干稳定度", f"{inhibition.torso_stability_degree:.2f}", "像素", "数值越小，躯干控制越稳定。"],
            ["抑制指标", "无效小动作频次", inhibition.invalid_motion_count, "次", "非规定动作的小幅异动次数。"],
            ["科研后测指标", "身体稳定性参数（CV）", f"{getattr(inhibition, 'body_stability_cv', 0.0):.6f}", "CV", "颈-髋连线动态偏移角度的变异系数。"],
            ["科研后测指标", "动作流畅性参数-过渡期加速度峰值", f"{getattr(inhibition, 'transition_acc_peak', 0.0):.6f}", "px/frame²", "越小越流畅。"],
            ["科研后测指标", "动作流畅性参数-过渡期急动度均值", f"{getattr(inhibition, 'transition_jerk_mean', 0.0):.6f}", "px/frame³", "越小越平顺。"],
            ["科研后测指标", "身体调控能力参数", f"{getattr(inhibition, 'body_control_ratio', 0.0):.6f}", "ratio", "过渡阶段位移幅度 / 定势阶段位移幅度。"],
        ]
    )

    for joint_name, stability in inhibition.joint_angle_stability.items():
        rows.append(["角度稳定性", joint_name, f"{stability:.2f}", "度", "角速度方差越小越平稳。"])

    rows.append(["", "", "", "", ""])
    rows.append(["关键点标准差明细", "关键点", "标准差", "像素", ""])
    for kp_name, std_value in sorted(inhibition.keypoint_std.items(), key=lambda x: x[1], reverse=True):
        rows.append(["关键点标准差", _KEYPOINT_LABELS.get(kp_name, kp_name), f"{std_value:.2f}", "像素", ""])

    _write_table_sheet(sheet, rows)


def _write_key_value_sheet(sheet, rows: list[tuple[str, object]]) -> None:
    header_fill = PatternFill("solid", fgColor="DCEBFA")
    header_font = Font(bold=True, color="16324D")
    for row_index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.cell(row=row_index, column=1).fill = header_fill
        sheet.cell(row=row_index, column=1).font = header_font
        sheet.cell(row=row_index, column=1).alignment = Alignment(vertical="center")
        sheet.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 78


def _write_table_sheet(sheet, rows: list[list[object]]) -> None:
    header_fill = PatternFill("solid", fgColor="DCEBFA")
    header_font = Font(bold=True, color="16324D")
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_index == 1:
                cell.fill = header_fill
                cell.font = header_font
    for column in sheet.columns:
        letter = column[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[letter].width = min(max(max_len + 4, 12), 44)


def _format_excel_metric_value(value, unit: str = "", signed: bool = False) -> object:
    if value is None:
        return ""
    numeric = round(float(value), 3)
    if unit:
        return f"{numeric:+.3f}{unit}" if signed else f"{numeric:.3f}{unit}"
    return f"{numeric:+.3f}" if signed else numeric


def _format_excel_seconds(value, signed: bool = False) -> object:
    if value is None:
        return ""
    numeric = round(float(value), 2)
    return f"{numeric:+.2f}s" if signed else f"{numeric:.2f}s"
